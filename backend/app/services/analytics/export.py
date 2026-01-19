"""Analytics export service for generating CSV and Excel reports"""

import csv
import io
import logging
from datetime import datetime
from decimal import Decimal
from enum import Enum
from typing import Any, Dict, List, Optional
from uuid import UUID

from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.deal import Deal
from app.models.bank_split import DealSplitRecipient
from app.models.dispute import Dispute
from app.models.user import User

logger = logging.getLogger(__name__)


class ExportFormat(str, Enum):
    """Supported export formats"""
    CSV = "csv"
    EXCEL = "xlsx"


class ExportType(str, Enum):
    """Types of exports available"""
    DEALS = "deals"
    PAYOUTS = "payouts"
    DISPUTES = "disputes"
    TIME_SERIES = "time_series"
    SUMMARY = "summary"


# Column definitions for each export type
DEALS_COLUMNS = [
    ("id", "ID сделки"),
    ("created_at", "Дата создания"),
    ("property_address", "Адрес объекта"),
    ("deal_type", "Тип сделки"),
    ("status", "Статус"),
    ("price", "Цена объекта"),
    ("commission_agent", "Комиссия агента"),
    ("client_name", "Клиент"),
    ("client_phone", "Телефон клиента"),
    ("agent_user_id", "ID агента"),
    ("payment_model", "Модель оплаты"),
]

PAYOUTS_COLUMNS = [
    ("id", "ID выплаты"),
    ("deal_id", "ID сделки"),
    ("created_at", "Дата создания"),
    ("role", "Роль"),
    ("inn", "ИНН"),
    ("calculated_amount", "Сумма"),
    ("payout_status", "Статус выплаты"),
    ("user_id", "ID пользователя"),
    ("organization_id", "ID организации"),
]

DISPUTES_COLUMNS = [
    ("id", "ID спора"),
    ("deal_id", "ID сделки"),
    ("created_at", "Дата создания"),
    ("initiator_user_id", "Инициатор"),
    ("reason", "Причина"),
    ("description", "Описание"),
    ("status", "Статус"),
    ("resolution", "Решение"),
    ("refund_requested", "Возврат запрошен"),
    ("refund_amount", "Сумма возврата"),
    ("refund_status", "Статус возврата"),
]

TIME_SERIES_COLUMNS = [
    ("date", "Дата"),
    ("deals_count", "Количество сделок"),
    ("volume", "Объём (руб)"),
    ("commission", "Комиссия (руб)"),
]

SUMMARY_COLUMNS = [
    ("metric", "Показатель"),
    ("value", "Значение"),
]

# Russian status translations
STATUS_TRANSLATIONS = {
    "draft": "Черновик",
    "awaiting_signatures": "Ожидание подписей",
    "signed": "Подписан",
    "invoiced": "Выставлен счёт",
    "payment_pending": "Ожидание оплаты",
    "payment_failed": "Ошибка оплаты",
    "hold_period": "Холд период",
    "payout_ready": "Готово к выплате",
    "payout_in_progress": "Выплата в процессе",
    "closed": "Закрыта",
    "cancelled": "Отменена",
    "dispute": "Спор",
    "refunded": "Возврат",
    "pending": "Ожидание",
    "paid": "Оплачено",
    "completed": "Завершено",
    "hold": "Холд",
    "open": "Открыт",
    "under_review": "На рассмотрении",
    "resolved": "Решён",
    "rejected": "Отклонён",
}

REASON_TRANSLATIONS = {
    "service_not_provided": "Услуга не оказана",
    "service_quality": "Качество услуги",
    "wrong_amount": "Неверная сумма",
    "fraud": "Мошенничество",
    "duplicate": "Дубликат",
    "other": "Другое",
}


def translate_status(status: str) -> str:
    """Translate status to Russian"""
    return STATUS_TRANSLATIONS.get(status, status)


def translate_reason(reason: str) -> str:
    """Translate reason to Russian"""
    return REASON_TRANSLATIONS.get(reason, reason)


def format_value(value: Any) -> str:
    """Format value for export"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, Decimal):
        return str(float(value))
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    return str(value)


class ExportService:
    """
    Service for exporting analytics data to various formats.

    Supports:
    - CSV export (built-in)
    - Excel export (requires openpyxl)
    """

    def __init__(self, db: AsyncSession):
        self.db = db

    async def export_deals(
        self,
        format: ExportFormat,
        user_id: Optional[int] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        status_filter: Optional[str] = None,
    ) -> bytes:
        """
        Export deals to CSV or Excel.

        Args:
            format: Export format (csv or xlsx)
            user_id: Filter by agent user ID
            start_date: Filter by start date
            end_date: Filter by end date
            status_filter: Filter by deal status

        Returns:
            Bytes of the exported file
        """
        # Build query
        query = select(Deal).where(Deal.payment_model == "bank_hold_split")

        if user_id:
            query = query.where(Deal.agent_user_id == user_id)
        if start_date:
            query = query.where(Deal.created_at >= start_date)
        if end_date:
            query = query.where(Deal.created_at <= end_date)
        if status_filter:
            query = query.where(Deal.status == status_filter)

        query = query.order_by(Deal.created_at.desc())

        result = await self.db.execute(query)
        deals = result.scalars().all()

        # Prepare rows
        rows = []
        for deal in deals:
            rows.append({
                "id": str(deal.id),
                "created_at": deal.created_at,
                "property_address": deal.property_address,
                "deal_type": deal.type,
                "status": translate_status(deal.status),
                "price": deal.price,
                "commission_agent": deal.commission_agent,
                "client_name": deal.client_name,
                "client_phone": deal.client_phone,
                "agent_user_id": deal.agent_user_id,
                "payment_model": deal.payment_model,
            })

        return self._generate_export(format, DEALS_COLUMNS, rows, "deals")

    async def export_payouts(
        self,
        format: ExportFormat,
        user_id: Optional[int] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        status_filter: Optional[str] = None,
    ) -> bytes:
        """Export payouts to CSV or Excel"""
        query = select(DealSplitRecipient)

        if user_id:
            query = query.where(DealSplitRecipient.user_id == user_id)
        if start_date:
            query = query.where(DealSplitRecipient.created_at >= start_date)
        if end_date:
            query = query.where(DealSplitRecipient.created_at <= end_date)
        if status_filter:
            query = query.where(DealSplitRecipient.payout_status == status_filter)

        query = query.order_by(DealSplitRecipient.created_at.desc())

        result = await self.db.execute(query)
        recipients = result.scalars().all()

        rows = []
        for r in recipients:
            rows.append({
                "id": str(r.id),
                "deal_id": str(r.deal_id),
                "created_at": r.created_at,
                "role": r.role,
                "inn": r.inn,
                "calculated_amount": r.calculated_amount,
                "payout_status": translate_status(r.payout_status) if r.payout_status else "",
                "user_id": r.user_id,
                "organization_id": str(r.organization_id) if r.organization_id else "",
            })

        return self._generate_export(format, PAYOUTS_COLUMNS, rows, "payouts")

    async def export_disputes(
        self,
        format: ExportFormat,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        status_filter: Optional[str] = None,
    ) -> bytes:
        """Export disputes to CSV or Excel (admin only)"""
        query = select(Dispute)

        if start_date:
            query = query.where(Dispute.created_at >= start_date)
        if end_date:
            query = query.where(Dispute.created_at <= end_date)
        if status_filter:
            query = query.where(Dispute.status == status_filter)

        query = query.order_by(Dispute.created_at.desc())

        result = await self.db.execute(query)
        disputes = result.scalars().all()

        rows = []
        for d in disputes:
            rows.append({
                "id": str(d.id),
                "deal_id": str(d.deal_id),
                "created_at": d.created_at,
                "initiator_user_id": d.initiator_user_id,
                "reason": translate_reason(d.reason) if d.reason else "",
                "description": d.description,
                "status": translate_status(d.status) if d.status else "",
                "resolution": d.resolution,
                "refund_requested": d.refund_requested,
                "refund_amount": d.refund_amount,
                "refund_status": translate_status(d.refund_status) if d.refund_status else "",
            })

        return self._generate_export(format, DISPUTES_COLUMNS, rows, "disputes")

    async def export_time_series(
        self,
        format: ExportFormat,
        days: int = 30,
        user_id: Optional[int] = None,
    ) -> bytes:
        """Export time series data to CSV or Excel"""
        from app.services.analytics.service import AnalyticsService

        analytics = AnalyticsService(self.db)
        series = await analytics.get_time_series(days=days, user_id=user_id)

        rows = []
        for point in series:
            rows.append({
                "date": point["date"],
                "deals_count": point["deals_count"],
                "volume": point["volume"],
                "commission": point["commission"],
            })

        return self._generate_export(format, TIME_SERIES_COLUMNS, rows, "time_series")

    async def export_summary(
        self,
        format: ExportFormat,
        user_id: Optional[int] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
    ) -> bytes:
        """Export summary statistics to CSV or Excel"""
        from app.services.analytics.service import AnalyticsService

        analytics = AnalyticsService(self.db)

        deal_stats = await analytics.get_deal_statistics(
            user_id=user_id,
            start_date=start_date,
            end_date=end_date,
        )

        payout_stats = await analytics.get_payout_statistics(
            user_id=user_id,
            start_date=start_date,
            end_date=end_date,
        )

        rows = [
            {"metric": "Всего сделок", "value": deal_stats["total_deals"]},
            {"metric": "Общий объём (руб)", "value": f"{deal_stats['total_volume']:,.2f}"},
            {"metric": "Общая комиссия (руб)", "value": f"{deal_stats['total_commission']:,.2f}"},
            {"metric": "Средний размер сделки (руб)", "value": f"{deal_stats['avg_deal_size']:,.2f}"},
            {"metric": "Средняя комиссия (руб)", "value": f"{deal_stats['avg_commission']:,.2f}"},
            {"metric": "", "value": ""},  # Empty row separator
            {"metric": "Ожидает выплаты (руб)", "value": f"{payout_stats['total_pending']:,.2f}"},
            {"metric": "Выплачено (руб)", "value": f"{payout_stats['total_paid']:,.2f}"},
            {"metric": "Количество получателей", "value": payout_stats["recipients_count"]},
        ]

        # Add status breakdown
        if deal_stats["deals_by_status"]:
            rows.append({"metric": "", "value": ""})
            rows.append({"metric": "=== Сделки по статусам ===", "value": ""})
            for status, count in deal_stats["deals_by_status"].items():
                rows.append({
                    "metric": translate_status(status),
                    "value": count,
                })

        return self._generate_export(format, SUMMARY_COLUMNS, rows, "summary")

    def _generate_export(
        self,
        format: ExportFormat,
        columns: List[tuple],
        rows: List[Dict],
        sheet_name: str,
    ) -> bytes:
        """Generate export file in specified format"""
        if format == ExportFormat.CSV:
            return self._generate_csv(columns, rows)
        elif format == ExportFormat.EXCEL:
            return self._generate_excel(columns, rows, sheet_name)
        else:
            raise ValueError(f"Unsupported export format: {format}")

    def _generate_csv(
        self,
        columns: List[tuple],
        rows: List[Dict],
    ) -> bytes:
        """Generate CSV file"""
        output = io.StringIO()

        # Write BOM for Excel UTF-8 compatibility
        output.write('\ufeff')

        writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

        # Write header
        writer.writerow([col[1] for col in columns])

        # Write data rows
        for row in rows:
            writer.writerow([
                format_value(row.get(col[0], ""))
                for col in columns
            ])

        return output.getvalue().encode('utf-8')

    def _generate_excel(
        self,
        columns: List[tuple],
        rows: List[Dict],
        sheet_name: str,
    ) -> bytes:
        """Generate Excel file (requires openpyxl)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return self._generate_csv(columns, rows)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name limit

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write header
        for col_idx, (col_key, col_name) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, (col_key, _) in enumerate(columns, 1):
                value = row.get(col_key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=format_value(value))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # Auto-adjust column widths
        for col_idx, (col_key, col_name) in enumerate(columns, 1):
            max_length = len(col_name)
            for row in rows:
                value = format_value(row.get(col_key, ""))
                max_length = max(max_length, len(str(value)))

            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


# Convenience function
async def export_analytics(
    db: AsyncSession,
    export_type: ExportType,
    format: ExportFormat,
    user_id: Optional[int] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    **kwargs,
) -> bytes:
    """
    Convenience function for exporting analytics.

    Args:
        db: Database session
        export_type: Type of export (deals, payouts, disputes, etc.)
        format: Export format (csv or xlsx)
        user_id: Optional user filter
        start_date: Optional start date filter
        end_date: Optional end date filter
        **kwargs: Additional arguments for specific export types

    Returns:
        Bytes of the exported file
    """
    service = ExportService(db)

    if export_type == ExportType.DEALS:
        return await service.export_deals(
            format=format,
            user_id=user_id,
            start_date=start_date,
            end_date=end_date,
            status_filter=kwargs.get("status_filter"),
        )
    elif export_type == ExportType.PAYOUTS:
        return await service.export_payouts(
            format=format,
            user_id=user_id,
            start_date=start_date,
            end_date=end_date,
            status_filter=kwargs.get("status_filter"),
        )
    elif export_type == ExportType.DISPUTES:
        return await service.export_disputes(
            format=format,
            start_date=start_date,
            end_date=end_date,
            status_filter=kwargs.get("status_filter"),
        )
    elif export_type == ExportType.TIME_SERIES:
        return await service.export_time_series(
            format=format,
            days=kwargs.get("days", 30),
            user_id=user_id,
        )
    elif export_type == ExportType.SUMMARY:
        return await service.export_summary(
            format=format,
            user_id=user_id,
            start_date=start_date,
            end_date=end_date,
        )
    else:
        raise ValueError(f"Unknown export type: {export_type}")
